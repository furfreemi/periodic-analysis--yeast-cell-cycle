import matlab.engine
import openpyxl
eng = matlab.engine.start_matlab()
wb = openpyxl.load_workbook('periodic_analysis2.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

wb_new = openpyxl.Workbook()
sheet_write = wb_new.active

permutations = [
[2, 4, 5, 1, 9, 6, 10, 8, 7, 3],
[1, 2, 4, 5, 6, 9, 10, 7, 8, 3],
[2, 1, 5, 4, 8, 3, 9, 10, 6, 7],
[4, 1, 2, 5, 3, 9, 6, 10, 8, 7]
]
p_count = 1
for p in permutations:
    i = 1
    while i <= 143:#143 = number of periodic functions as according to data
        genename = sheet.cell(row=i, column=1).value
        sheet_write.cell(row=i, column=1).value = genename
        data = []
        for val in p:
            j = val + 3
            data.append(float(sheet.cell(row=i, column=j).value))
        #call matlab function: current data order=data, current permutation=p
        periodic_level = eng.compute_periodicity(data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9])

        sheet_write.cell(row=i, column=(p_count + 1)).value = periodic_level

        i = i + 1
    p_count = p_count + 1

wb_new.save('results_heatmap.xlsx')
