import requests
import openpyxl

wb_new = openpyxl.Workbook()
sheet_write = wb_new.active

wb = openpyxl.load_workbook('scrambled.xlsx')
sheet = wb.get_sheet_by_name('scrambled.csv')

i = 1
periodiccount = 0

for gene in sheet.columns[0]:
    genename = gene.value
    print genename
    page = requests.get('http://webapps.fhcrc.org/labs/noble/cellcycle/getOrf.php?orf=' + genename)
    #print page.content
    periodic = 'false'
    if page.content.find('not considered periodic') == -1:
        periodic = 'true'
        periodiccount = periodiccount + 1
    #else:
    #    print 'not periodic'
    rank = 0
    if page.content.find('ranks this transcript <b>') != -1:
        s1 = page.content.split('ranks this transcript <b>')
        if len(s1) > 0:
            s2 = s1[1].split('</b> among all')
            print 'rank: ' + s2[0]
            rank = s2[0]
    values = []
    j = 2
    while j <= 11:
        values.append(sheet.cell(row=i, column=j).value)
        j = j + 1
    print values
    print ""

    sheet_write.cell(row=i, column=1).value = genename
    sheet_write.cell(row=i, column=2).value = periodic
    sheet_write.cell(row=i, column=3).value = rank
    k = 4
    for val in values:
        sheet_write.cell(row=i, column=k).value = val
        k = k + 1
    i = i + 1
print 'total periodic: ' + str(periodiccount)

wb_new.save('periodic_analysis.xlsx')
