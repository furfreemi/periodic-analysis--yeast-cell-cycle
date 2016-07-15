import openpyxl

wb = openpyxl.load_workbook('results_heatmap.xlsx')
sheet = wb.active
permutation = 2

while permutation <= 4:
    gene = 1
    while gene <= 143:
        if sheet.cell(row=gene, column=7).value < sheet.cell(row=gene, column=(permutation + 1)).value:
            sheet.cell(row=gene, column=7).value = sheet.cell(row=gene, column=(permutation + 1)).value
            sheet.cell(row=gene, column=6).value = permutation
        print sheet.cell(row=gene, column = 7).value
        gene = gene + 1
    permutation = permutation + 1
